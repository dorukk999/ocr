import streamlit as st
import pandas as pd
from google import genai
from google.genai import types
from google.genai import errors as genai_errors
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
import json
import base64
import concurrent.futures
import difflib
import hashlib
import re
import time
import uuid
import zipfile
from PIL import Image
import io

# --- SETTINGS ---
APP_PASSWORD = "DataOCR_98123606513" # You can update this password

# Page Configuration
st.set_page_config(page_title="AI Document Extraction Tool", layout="wide")

# Initialize Session State
if "authenticated" not in st.session_state: st.session_state["authenticated"] = False
if "api_key" not in st.session_state: st.session_state["api_key"] = ""
if "uploader_key" not in st.session_state: st.session_state["uploader_key"] = 0

# Master Corporate Schema
OFFICIAL_SCHEMA_ORDER = [
    "Document Type", "Full Name", "Arabic Name", "Nationality",
    "Emirates ID Number (784-xxxx-xxxxxxx-x)", "Passport Number",
    "Work Permit Number (9 Digits)", "Personal Number (14 Digits)",
    "Permit / Security Pass / Card Number", "UID / Unified Number",
    "Date of Birth", "Date of Issue", "Date of Expiry", 
    "Employer / Company Name", "Occupation / Trade", "Site Location", 
    "Arabic Site Location Names", "English Translation of Arabic Site Locations", 
    "Issuing Authority", "Any other visible document-specific fields", 
    "Confidence level for each extracted field", "Remarks for unclear or doubtful fields"
]

MAX_RETRIES = 3

# Fields the model has been observed to confuse with one another; give it explicit, disambiguating rules.
FIELD_HINTS = {
    "Document Type": "A short description of the specific document that always includes one of these category words, so it can be reliably classified downstream: 'Identity' or 'Emirates ID' for a Resident Identity Card; 'Passport' for a passport; 'Visa' or 'Residence' or 'Entry Permit' for a UAE residence visa / entry permit (even if its own header just says 'Residence'); 'Labor Card' or 'Work Permit' for a labor/work permit card; 'Security Pass' or 'Defense' or 'Military' for a Ministry of Defense security/access pass. A Labor Card has no on-card title of its own (it never literally prints 'Labor Card' anywhere) — recognize it instead by its 'UNITED ARAB EMIRATES MINISTRY OF HUMAN RESOURCES & EMIRATISATION' header combined with 'Work Permit NO' and 'Personal NO' fields, and still label it 'Labor Card' / 'Work Permit' as the Document Type.",
    "Full Name": "the document holder's own given name(s) and surname, exactly as printed near a 'Name' / 'الاسم' label. Never the nationality, country, employer, or job title. Always output this in Latin/English script — if the document only shows the name in Arabic, phonetically transliterate it to Latin letters. Never put Arabic script in this field (Arabic script belongs only in 'Arabic Name'). Always order it as [Given Name(s)] [Surname] (e.g. 'Kali Bahadur Thapa'), never [Surname] [Given Name(s)] — even if a passport's own layout lists the surname field first, reorder it to given-name-first so the same person's name is formatted identically across every document type.",
    "Arabic Name": "the same person's name in Arabic script, next to the Arabic 'الاسم' label. Never the Arabic word for a country/nationality.",
    "Nationality": "the person's country/nationality, next to a 'Nationality' / 'الجنسية' label (e.g. Nepal, India). Never the person's name.",
    "Date of Birth": "the document holder's date of birth. On an Emirates ID or Passport that has a machine-readable zone (MRZ — the row(s) of monospaced characters and '<' filler symbols at the bottom/back), the MRZ encodes this same date as a 6-digit YYMMDD string. Before finalizing this field, decode that MRZ date and cross-check it against the printed Date of Birth you extracted — they describe the same day, so if they disagree, re-read both and correct whichever was misread rather than reporting the printed value blindly.",
    "Date of Expiry": "the document's expiry date. On an Emirates ID or Passport that has a machine-readable zone (MRZ — the row(s) of monospaced characters and '<' filler symbols at the bottom/back), the MRZ encodes this same date as a 6-digit YYMMDD string. Before finalizing this field, decode that MRZ date and cross-check it against the printed Date of Expiry you extracted — they describe the same day, so if they disagree, re-read both and correct whichever was misread rather than reporting the printed value blindly.",
    "Emirates ID Number (784-xxxx-xxxxxxx-x)": "the person's Emirates ID number in the 784-YYYY-NNNNNNN-C format (15 digits total), next to an 'ID Number' / 'رقم الهوية' label. This can appear not only on the Emirates ID card itself but also on a Residence Visa / Entry Permit and occasionally a Security Pass — always capture it wherever this exact 15-digit, 784-prefixed, explicitly labeled number appears, regardless of document type. Never confuse it with the Passport Number, a Visa File Number, or any other number on the document.",
    "UID / Unified Number": "ONLY fill this in if the document has an explicit label reading 'Unified No.', 'UID', 'Unified Number', or the Arabic 'الرقم الموحد' directly next to the value. This field is an exception to the contextual-inference rule below: never derive, compute, or extract it from the Emirates ID number, card number, chip number, or any part of the MRZ string, even if the digit count matches — those are different fields, not the Unified Number. If there is no such explicit label, always output 'N/A' here, no matter how plausible a nearby number looks.",
    "Issuing Authority": "the name of the organization/government body that issued the document (e.g. 'Federal Authority for Identity & Citizenship, Customs & Port Security', 'MOFA, Department of Passport', 'UAE Ministry of Defense', 'Ministry of Human Resources & Emiratisation') — usually found in the document's header/logo area. This is NEVER a city or place name like 'Al Ain' — a city where the document was issued is the 'Issuing Place' and belongs only in 'Any other visible document-specific fields', not here.",
    "Personal Number (14 Digits)": "This field is an exception to the contextual-inference rule below: ONLY fill it in if the document has an explicit label reading 'Personal Number', 'Personal NO', or the Arabic 'الرقم الشخصي' directly next to the value, AND that labeled value is exactly 14 digits long. This field is most reliably found, correctly labeled, on Labor Card / Work Permit documents. If a 'Personal No.' or similar label exists but its value is NOT exactly 14 digits (e.g. a 10 or 11-digit Personal No. on a passport), do NOT place it in this field — output 'N/A' here instead, and record the label and its actual value under 'Any other visible document-specific fields'. Never fill this field in just because some unrelated number on the document happens to be 14 digits long without that explicit label, and never round, pad, or trim a value to force it to 14 digits.",
    "Work Permit Number (9 Digits)": "This field is an exception to the contextual-inference rule below: ONLY fill it in if the document has an explicit label reading 'Work Permit Number', 'Work Permit NO', or the Arabic equivalent directly next to the value, AND that labeled value is exactly 9 digits long. This field is most reliably found, correctly labeled, on Labor Card / Work Permit documents. On a Labor Card, the 'Work Permit NO' cell often also contains a small unrelated place name (e.g. 'العين' / 'Al Ain') printed next to or before the number — that place name is NOT part of the Work Permit Number and must be excluded; extract only the 9-digit number itself. If a similarly labeled value exists but is not exactly 9 digits, output 'N/A' here instead, and record the label and its actual value under 'Any other visible document-specific fields'. Never fill this field in just because some unrelated number happens to be 9 digits long without that explicit label, and never round, pad, or trim a value to force it to 9 digits.",
    "Permit / Security Pass / Card Number": "the document's own primary reference/serial number — NOT the Emirates ID Number, Passport Number, Work Permit Number, or UID / Unified Number. On a Security Pass this is its card/pass number. On a Residence Visa / Entry Permit this is the immigration 'File' number (format like '152/2026/2/13668'), next to a 'File' / 'الملف' label — use this field for that File number. Never confuse this field with the Emirates ID Number (784-xxxx-xxxxxxx-x format) or the Passport Number even when they appear on the same document.",
}

# Common nationality/demonym words seen on UAE labor/ID documents — used to flag a Full Name
# that looks like it was actually swapped with the Nationality field.
NATIONALITY_WORDS = {
    "nepal", "nepali", "nepalese", "india", "indian", "pakistan", "pakistani",
    "bangladesh", "bangladeshi", "philippines", "filipino", "filipina",
    "sri lanka", "sri lankan", "egypt", "egyptian", "sudan", "sudanese",
    "syria", "syrian", "jordan", "jordanian", "yemen", "yemeni", "lebanon",
    "lebanese", "iraq", "iraqi", "iran", "iranian", "afghanistan", "afghan",
    "ethiopia", "ethiopian", "kenya", "kenyan", "uganda", "ugandan", "somalia",
    "somali", "morocco", "moroccan", "tunisia", "tunisian", "indonesia",
    "indonesian", "bhutan", "bhutanese", "myanmar", "burmese", "china",
    "chinese", "uae", "emirati", "saudi", "saudi arabian", "oman", "omani",
    "kuwait", "kuwaiti", "qatar", "qatari", "bahrain", "bahraini", "britain",
    "british", "america", "american", "canada", "canadian", "australia",
    "australian", "nigeria", "nigerian", "ghana", "ghanaian", "south africa",
    "south african",
}

COLUMN_WIDTHS = {
    "Source_File_Name": 18, "Document Type": 20, "Full Name": 22,
    "Reference ID": 16, "Missing Documents": 26, "Name Consistency Flag": 40,
    "Arabic Name": 22, "Nationality": 14,
    "Emirates ID Number (784-xxxx-xxxxxxx-x)": 24, "Passport Number": 16,
    "Work Permit Number (9 Digits)": 16, "Personal Number (14 Digits)": 16,
    "Permit / Security Pass / Card Number": 20, "UID / Unified Number": 18,
    "Date of Birth": 14, "Date of Issue": 14, "Date of Expiry": 14,
    "Employer / Company Name": 32, "Occupation / Trade": 18,
    "Site Location": 30, "Arabic Site Location Names": 30,
    "English Translation of Arabic Site Locations": 32, "Issuing Authority": 30,
    "Any other visible document-specific fields": 40,
    "Confidence level for each extracted field": 40,
    "Remarks for unclear or doubtful fields": 40, "Review Flags": 30,
    "Processing Time (s)": 12,
}

ARABIC_SCRIPT_RE = re.compile(r'[؀-ۿ]')

def compute_review_flags(safe_json):
    full_name = str(safe_json.get("Full Name", "")).strip()
    flags = []
    if full_name and full_name.lower() not in ("-", "n/a"):
        if ARABIC_SCRIPT_RE.search(full_name):
            flags.append("Full Name contains Arabic script — verify")
        elif difflib.get_close_matches(full_name.lower(), NATIONALITY_WORDS, n=1, cutoff=0.8):
            flags.append("Full Name looks like a nationality — verify")
        elif " " not in full_name:
            flags.append("Full Name is a single word — verify")

    uid = str(safe_json.get("UID / Unified Number", "")).strip()
    emirates_id = str(safe_json.get("Emirates ID Number (784-xxxx-xxxxxxx-x)", "")).strip()
    if uid and uid.lower() not in ("-", "n/a"):
        uid_digits = re.sub(r'\D', '', uid)
        id_digits = re.sub(r'\D', '', emirates_id)
        if uid_digits and id_digits:
            match = difflib.SequenceMatcher(None, uid_digits, id_digits).find_longest_match(0, len(uid_digits), 0, len(id_digits))
            fully_contained = match.size == len(uid_digits) and len(uid_digits) >= 5
            if match.size >= 8 or fully_contained:
                flags.append("UID overlaps with Emirates ID/MRZ digits, likely derived not labeled — verify")

    return "; ".join(flags)

# Buckets a document falls into, used to prefix its columns in the per-person wide view.
# Order matters: more specific keywords are checked before generic ones.
DOC_TYPE_BUCKETS = [
    ("Emirates ID", ["resident identity", "emirates id", " eid", "identity card"]),
    ("Passport", ["passport"]),
    ("Visa", ["visa", "entry permit", "residence permit", "residence", "residency"]),
    ("Labor Card", ["labor card", "labour card", "labor permit", "labour permit", "work permit", "human resources", "emiratisation", "mohre"]),
    ("Security Pass", ["security pass", "defense", "ministry of defense", "military", "access card"]),
]
EXPECTED_DOC_TYPES = [name for name, _ in DOC_TYPE_BUCKETS]

def classify_doc_type(document_type_text, filename):
    text = f"{document_type_text} {filename}".lower()
    for bucket_name, keywords in DOC_TYPE_BUCKETS:
        if any(kw in text for kw in keywords):
            return bucket_name
    return "Other Document"

def extract_reference_id(filename):
    """Pulls the employee reference number from a filename like '1024_EID.jpg'
    (everything before the first separator). Falls back to the full filename
    stem if no separator is found."""
    stem = str(filename).rsplit(".", 1)[0]
    for sep in ("_", "-", " "):
        if sep in stem:
            return stem.split(sep, 1)[0].strip()
    return stem.strip()

SUPPORTED_EXTENSIONS = ("png", "jpg", "jpeg", "pdf")

def extract_zip_files(zip_bytes):
    """Reads a ZIP where each employee's documents live in their own folder
    (e.g. '09592813/EID.jpg'). Returns (file_bytes, synthesized_filename) pairs
    with the folder name prefixed onto the filename (e.g. '09592813_EID.jpg'),
    so the existing reference-ID grouping picks it up with no renaming needed."""
    results = []
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        for info in zf.infolist():
            if info.is_dir():
                continue
            name = info.filename
            if name.startswith("__MACOSX/") or name.split("/")[-1].startswith("."):
                continue
            ext = name.rsplit(".", 1)[-1].lower() if "." in name else ""
            if ext not in SUPPORTED_EXTENSIONS:
                continue
            parts = [p for p in name.split("/") if p]
            if len(parts) >= 2:
                synthesized_name = f"{parts[-2]}_{parts[-1]}"
            else:
                synthesized_name = parts[-1] if parts else name
            results.append((zf.read(info), synthesized_name))
    return results

def limit_by_reference(extracted, limit):
    """Keeps every file belonging to the first `limit` distinct reference IDs
    (i.e. the first N employee folders) rather than an arbitrary file-count
    cutoff, so a test run always contains complete document sets instead of
    splitting a folder in half."""
    if limit <= 0:
        return extracted
    accepted_refs = []
    kept = []
    for fb, fn in extracted:
        ref = extract_reference_id(fn)
        if ref not in accepted_refs:
            if len(accepted_refs) >= limit:
                continue
            accepted_refs.append(ref)
        kept.append((fb, fn))
    return kept

WORD_MISMATCH_RATIO_THRESHOLD = 0.8
OVERALL_MISMATCH_RATIO_THRESHOLD = 0.5

def _normalize_name(name):
    return re.sub(r'\s+', ' ', str(name).strip().lower())

def compute_name_consistency_flag(names_by_bucket):
    """names_by_bucket: dict of {bucket: Full Name string} for one employee's
    documents. Flags when two documents disagree on the person's name by more
    than simple casing/spacing/spelling-variant differences — this is a
    stronger signal than picking the first non-empty name, since a bad read on
    one document would otherwise be silently discarded once a good name from
    another document fills the shared 'Full Name' field first.
    When both names have the same word count, compares word-by-word (catches
    e.g. one wrong letter in a single given/middle name, which a whole-string
    ratio would dilute and miss). When word counts differ — very common and
    usually fine, since an Emirates ID often includes a father's name that a
    passport omits — falls back to a lenient whole-string ratio so a merely
    longer/shorter but otherwise matching name isn't flagged."""
    entries = [(b, str(n).strip()) for b, n in names_by_bucket.items() if str(n).strip() and str(n).strip().lower() not in ("-", "n/a")]
    mismatches = []
    for i in range(len(entries)):
        for j in range(i + 1, len(entries)):
            bucket_a, name_a = entries[i]
            bucket_b, name_b = entries[j]
            tokens_a = _normalize_name(name_a).split(' ')
            tokens_b = _normalize_name(name_b).split(' ')
            if len(tokens_a) == len(tokens_b):
                word_ratios = [difflib.SequenceMatcher(None, ta, tb).ratio() for ta, tb in zip(tokens_a, tokens_b)]
                is_mismatch = min(word_ratios) < WORD_MISMATCH_RATIO_THRESHOLD
            else:
                overall_ratio = difflib.SequenceMatcher(None, _normalize_name(name_a), _normalize_name(name_b)).ratio()
                is_mismatch = overall_ratio < OVERALL_MISMATCH_RATIO_THRESHOLD
            if is_mismatch:
                mismatches.append(f"{bucket_a}: '{name_a}' vs {bucket_b}: '{name_b}'")
    return "; ".join(mismatches)

def build_person_view(df):
    """One row per employee, grouped by the reference number prefixed on each
    filename (e.g. '1024_EID.jpg' -> '1024') rather than the OCR'd Full Name,
    since the same person's name can read differently across document types.
    Each document's fields are prefixed by document type (e.g. 'Passport -
    Passport Number'). Also reports which of the 5 expected document types
    are missing for each employee, and flags employees whose Full Name reads
    meaningfully differently across their own documents (worth a manual check,
    since the merged 'Full Name' column only keeps the first non-N/A reading
    and would otherwise hide a bad OCR read on one of the other documents)."""
    df = df.copy()
    df["__bucket"] = [classify_doc_type(dt, fn) for dt, fn in zip(df.get("Document Type", ""), df.get("Source_File_Name", ""))]
    df["__ref"] = [extract_reference_id(fn) for fn in df.get("Source_File_Name", "")]

    shared_cols = ["Full Name", "Arabic Name", "Nationality"]
    per_doc_cols = [c for c in df.columns if c not in shared_cols + ["__bucket", "__ref", "Source_File_Name"]]

    people = {}
    order = []
    names_by_ref = {}
    for _, row in df.iterrows():
        ref = row["__ref"] or f"__unref__{row.get('Source_File_Name', '')}"
        if ref not in people:
            people[ref] = {"Reference ID": ref}
            order.append(ref)
            names_by_ref[ref] = {}
        for c in shared_cols:
            if not str(people[ref].get(c, "")).strip() or str(people[ref][c]).strip().lower() in ("-", "n/a"):
                people[ref][c] = row.get(c, "")
        bucket = row["__bucket"]
        names_by_ref[ref][bucket] = row.get("Full Name", "")
        for c in per_doc_cols:
            col_name = f"{bucket} - {c}"
            value = row.get(c, "")
            existing = people[ref].get(col_name)
            if existing is None or str(existing).strip().lower() in ("", "-", "n/a"):
                people[ref][col_name] = value

    for ref in order:
        missing = [b for b in EXPECTED_DOC_TYPES if str(people[ref].get(f"{b} - Document Type", "")).strip().lower() in ("", "-", "n/a")]
        people[ref]["Missing Documents"] = ", ".join(missing) if missing else "None"
        people[ref]["Name Consistency Flag"] = compute_name_consistency_flag(names_by_ref[ref])

    front_cols = ["Reference ID", "Full Name", "Missing Documents", "Name Consistency Flag", "Arabic Name", "Nationality"]
    result = pd.DataFrame([people[k] for k in order])
    ordered_cols = front_cols + [c for c in result.columns if c not in front_cols]
    return result[ordered_cols]

def build_excel_bytes(sheets):
    """sheets: dict of {sheet_name: DataFrame}. Writes one formatted sheet per entry."""
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            safe_name = sheet_name[:31]
            df.to_excel(writer, index=False, sheet_name=safe_name)
            worksheet = writer.sheets[safe_name]
            wrap = Alignment(wrap_text=True, vertical="top")
            for col_idx, col_name in enumerate(df.columns, start=1):
                base_name = col_name.split(" - ", 1)[-1]
                width = COLUMN_WIDTHS.get(col_name, COLUMN_WIDTHS.get(base_name, 18))
                worksheet.column_dimensions[get_column_letter(col_idx)].width = width
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = wrap
            for i in range(2, worksheet.max_row + 1):
                worksheet.row_dimensions[i].height = 60
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
            worksheet.freeze_panes = "A2"
    return buffer.getvalue()

def optimize_image(file_bytes):
    img = Image.open(io.BytesIO(file_bytes))
    if img.mode in ("RGBA", "P"): img = img.convert("RGB")
    max_size = 1600
    if img.width > max_size or img.height > max_size:
        img.thumbnail((max_size, max_size), Image.Resampling.LANCZOS)
    buffer = io.BytesIO()
    img.save(buffer, format="JPEG", quality=90)
    return buffer.getvalue()

def process_file_backend(file_bytes, unique_filename, incoming_key, target_schema, target_model):
    start_time = time.time()
    try:
        client = genai.Client(api_key=incoming_key)
        is_pdf = unique_filename.lower().endswith('.pdf')
        if is_pdf:
            base64_data = base64.b64encode(file_bytes).decode("utf-8")
            mime_type = "application/pdf"
        else:
            optimized_bytes = optimize_image(file_bytes)
            base64_data = base64.b64encode(optimized_bytes).decode("utf-8")
            mime_type = "image/jpeg"

        field_rules = "\n".join(f"- {k}: {v}" for k, v in FIELD_HINTS.items() if k in target_schema)
        schema_prompt = (
            "You are an advanced corporate OCR engine. "
            f"Output exactly these keys: {json.dumps(target_schema)}. "
            "Every value must be a plain string (or number/date as text) — never a nested object or dict. "
            "Do not attach per-field confidence or remarks inside a value; use the dedicated "
            "'Confidence level for each extracted field' and 'Remarks for unclear or doubtful fields' keys for that instead.\n"
            f"Field-specific rules:\n{field_rules}\n"
            "Never substitute a DIFFERENT field's value into a field it doesn't belong to just because it "
            "seems like the closest fit — for example, a passport's Citizenship No. is not a "
            "'UID / Unified Number', and a Place of Birth is not a 'Site Location'. If no genuine counterpart "
            "for a field exists on the document, output 'N/A' for it and put that unrelated leftover value in "
            "'Any other visible document-specific fields' instead.\n"
            "This does NOT mean you should leave a field as 'N/A' when the correct value for THAT SAME field "
            "is clearly identifiable from context or position even without a printed label — e.g. if a "
            "country/nationality word appears where nationality is normally shown, still fill in 'Nationality' "
            "with it, and just note in 'Remarks' that it wasn't explicitly labeled."
        )

        last_error = None
        for attempt in range(MAX_RETRIES):
            try:
                response = client.models.generate_content(
                    model=target_model,
                    contents=[schema_prompt, {"inline_data": {"data": base64_data, "mime_type": mime_type}}],
                    config=types.GenerateContentConfig(response_mime_type="application/json")
                )
                extracted_json = json.loads(response.text.strip())
                safe_json = {"Source_File_Name": unique_filename}
                for col in target_schema:
                    found_key = next((k for k in extracted_json if k.lower().strip() == col.lower().strip()), None)
                    value = extracted_json[found_key] if found_key else "-"
                    if isinstance(value, dict):
                        value = value.get("value", value)
                    safe_json[col] = value
                safe_json["Review Flags"] = compute_review_flags(safe_json)
                safe_json["Processing Time (s)"] = round(time.time() - start_time, 1)
                return safe_json
            except genai_errors.APIError as e:
                last_error = e
                is_retryable = e.code in (429, 500, 503)
                if is_retryable and attempt < MAX_RETRIES - 1:
                    time.sleep(2 ** attempt)
                    continue
                raise
        raise last_error
    except Exception as e:
        return {
            "Source_File_Name": unique_filename,
            "Document Type": "FAILED",
            "Remarks for unclear or doubtful fields": str(e),
            "Processing Time (s)": round(time.time() - start_time, 1),
        }

CHUNK_SIZE = 50
MAX_WORKERS = 5

def file_signature(file_bytes):
    return hashlib.sha1(file_bytes).hexdigest()

def process_batch(files_data, api_key, schema, model, progress_bar, table_placeholder, total):
    """files_data: list of (bytes, display_filename, unique_key, signature). unique_key
    (not the filename) is used to track failures/retries, since the same
    filename (e.g. 'EID.jpg') commonly repeats across different people. signature
    is a content hash used to skip files already processed in this session."""
    processed_count = 0
    last_render = 0.0
    for i in range(0, len(files_data), CHUNK_SIZE):
        chunk = files_data[i:i + CHUNK_SIZE]
        chunk_lookup = {uk: (fb, sig) for fb, fn, uk, sig in chunk}
        with concurrent.futures.ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
            futures = {executor.submit(process_file_backend, fb, fn, api_key, schema, model): uk for fb, fn, uk, sig in chunk}
            for future in concurrent.futures.as_completed(futures):
                unique_key = futures[future]
                result = future.result()
                result["__key"] = unique_key
                st.session_state["batch_results"].append(result)
                processed_count += 1
                file_bytes, sig = chunk_lookup[unique_key]
                if result.get("Document Type") == "FAILED":
                    st.session_state["failed_bytes"][unique_key] = (result["Source_File_Name"], file_bytes, sig)
                else:
                    st.session_state["failed_bytes"].pop(unique_key, None)
                    st.session_state["processed_signatures"].add(sig)
                progress_bar.progress(processed_count / total)
                now = time.time()
                if now - last_render > 0.5 or processed_count == total:
                    display_df = pd.DataFrame(st.session_state["batch_results"]).drop(columns="__key", errors="ignore")
                    table_placeholder.dataframe(display_df, use_container_width=True)
                    last_render = now

def main():
    st.title("📂 Corporate AI Data Extraction Tool")

    # --- 1. STEP: APP LOGIN ---
    if not st.session_state["authenticated"]:
        st.subheader("🔒 Enter Application Password")
        password_input = st.text_input("Password:", type="password")
        if st.button("Login"):
            if password_input == APP_PASSWORD:
                st.session_state["authenticated"] = True
                st.rerun()
            else:
                st.error("Incorrect password!")
        return

    # --- 2. STEP: API SETTINGS & MAIN OPERATIONS ---
    with st.sidebar:
        st.header("⚙️ Settings")
        st.session_state["api_key"] = st.text_input("Enter your Gemini API Key:", type="password", value=st.session_state["api_key"])
        selected_model = st.selectbox(
            "Model",
            ["models/gemini-2.5-flash", "models/gemini-2.5-flash-lite"],
            index=0,
            help="flash-lite is faster and cheaper; test accuracy on your documents before relying on it."
        )
        custom_field = st.text_input("➕ Extract Custom Field")
        enable_consolidation = st.checkbox(
            "Enable Person-Based Consolidation",
            help="One row per employee instead of one row per document. Requires files to be named with a shared reference number prefix, e.g. '1024_EID.jpg', '1024_Passport.jpg', '1024_Visa.jpg', '1024_LaborCard.jpg', '1024_SecurityPass.jpg' — grouping uses this number, not the OCR'd name. Each document's fields get their own prefixed columns, and a 'Missing Documents' report sheet lists which of the 5 expected documents (Emirates ID, Passport, Visa, Labor Card, Security Pass) are missing per employee."
        )
        if st.button("Logout"):
            st.session_state["authenticated"] = False
            st.rerun()

    active_schema = OFFICIAL_SCHEMA_ORDER.copy()
    if custom_field: active_schema.insert(19, custom_field)

    if "batch_results" not in st.session_state: st.session_state["batch_results"] = []
    if "failed_bytes" not in st.session_state: st.session_state["failed_bytes"] = {}
    if "processed_signatures" not in st.session_state: st.session_state["processed_signatures"] = set()

    uploaded_files = st.file_uploader("Upload Documents", type=["png", "jpg", "jpeg", "pdf"], accept_multiple_files=True)

    if st.session_state["batch_results"] and st.button("🗑️ Clear All Results (start a new batch)"):
        st.session_state["batch_results"] = []
        st.session_state["failed_bytes"] = {}
        st.session_state["processed_signatures"] = set()
        st.rerun()

    if uploaded_files and st.session_state["api_key"] and st.button("🚀 Run Extraction Pipeline"):
        files_data = []
        skipped = 0
        for f in uploaded_files:
            fb = f.read()
            sig = file_signature(fb)
            if sig in st.session_state["processed_signatures"]:
                skipped += 1
                continue
            files_data.append((fb, f.name, f"{f.name}::{uuid.uuid4().hex[:8]}", sig))
        if skipped:
            st.info(f"⏭️ Skipped {skipped} file(s) already processed earlier in this session (same content re-selected).")
        if files_data:
            total = len(files_data)
            progress_bar = st.progress(0)
            table_placeholder = st.empty()
            process_batch(files_data, st.session_state["api_key"], active_schema, selected_model, progress_bar, table_placeholder, total)
            st.success(f"🎉 Processing completed — {len(st.session_state['batch_results'])} total file(s) accumulated so far.")

    st.markdown("**Or upload a ZIP** (one folder per employee, e.g. `09592813/EID.jpg`, `09592813/Passport.jpg` — the folder name becomes the reference ID automatically, no renaming needed):")
    zip_file = st.file_uploader("Upload ZIP", type=["zip"], label_visibility="collapsed")
    zip_test_limit = st.number_input(
        "Test with first N employees only (0 = process all)",
        min_value=0, value=0, step=5,
        help="Counted by folder, not by file — e.g. 10 means all documents for the first 10 employee folders, never a partial folder. Use this to try a small sample before committing to the full batch. Files already processed in a test run won't be reprocessed when you later run the rest."
    )

    if zip_file and st.session_state["api_key"] and st.button("📦 Process ZIP"):
        extracted = extract_zip_files(zip_file.read())
        if zip_test_limit > 0:
            extracted = limit_by_reference(extracted, zip_test_limit)
        files_data = []
        skipped = 0
        for fb, fn in extracted:
            sig = file_signature(fb)
            if sig in st.session_state["processed_signatures"]:
                skipped += 1
                continue
            files_data.append((fb, fn, f"{fn}::{uuid.uuid4().hex[:8]}", sig))
        if skipped:
            st.info(f"⏭️ Skipped {skipped} file(s) already processed earlier in this session.")
        if files_data:
            total = len(files_data)
            progress_bar = st.progress(0)
            table_placeholder = st.empty()
            process_batch(files_data, st.session_state["api_key"], active_schema, selected_model, progress_bar, table_placeholder, total)
            st.success(f"🎉 ZIP processed — {len(st.session_state['batch_results'])} total file(s) accumulated so far.")
        else:
            st.warning("No new supported files (png/jpg/jpeg/pdf) found in the ZIP.")

    if st.session_state["failed_bytes"]:
        n_failed = len(st.session_state["failed_bytes"])
        st.warning(f"⚠️ {n_failed} file(s) failed after retries.")
        if st.button(f"🔁 Retry {n_failed} Failed File(s) Only"):
            failed_keys = set(st.session_state["failed_bytes"].keys())
            st.session_state["batch_results"] = [r for r in st.session_state["batch_results"] if r.get("__key") not in failed_keys]
            files_data = [(fb, fn, uk, sig) for uk, (fn, fb, sig) in st.session_state["failed_bytes"].items()]
            total = len(files_data)
            progress_bar = st.progress(0)
            table_placeholder = st.empty()
            process_batch(files_data, st.session_state["api_key"], active_schema, selected_model, progress_bar, table_placeholder, total)
            st.rerun()

    if st.session_state["batch_results"]:
        df = pd.DataFrame(st.session_state["batch_results"]).drop(columns="__key", errors="ignore")
        if enable_consolidation and "Full Name" in df.columns:
            df = build_person_view(df)
            missing_df = df[["Reference ID", "Full Name", "Missing Documents", "Name Consistency Flag"]]
            sheets = {"Report": df, "Missing Documents": missing_df}
        else:
            sheets = {"Report": df}

        st.download_button(
            "📊 Download Excel Report (.xlsx)",
            build_excel_bytes(sheets),
            "Consolidated_Report.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        st.download_button("📥 Download CSV Report", df.to_csv(index=False).encode('utf-8-sig'), "Consolidated_Report.csv", "text/csv")

if __name__ == "__main__":
    main()
